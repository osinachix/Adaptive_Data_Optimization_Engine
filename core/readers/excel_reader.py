import itertools
from collections.abc import Iterator
from pathlib import Path

import openpyxl
import polars as pl

from core.types import Schema

# Excel has no lazy scanner in Polars or PyArrow (both load a whole sheet
# at once); the schema is inferred from a bounded prefix of rows, not the
# whole sheet, and read_batches() streams the rest from openpyxl's
# read-only row iterator (backed by incremental XML parsing, not a fully
# materialized worksheet), so this still honors invariant I1.
_SCHEMA_SAMPLE_ROWS = 1_000


class ExcelReader:
    """Streams the first worksheet of an .xlsx file as row batches.

    Uses openpyxl in read_only mode, whose iter_rows() is a lazy,
    incrementally-parsed row iterator rather than a fully materialized
    worksheet - the only Excel-reading approach here that avoids loading
    the whole sheet into memory at once. Only the first (active)
    worksheet is read; multi-sheet workbooks are out of scope. Formula
    cells are read as their last-computed value (data_only=True), not the
    formula text - if a workbook was never opened/saved in a real
    spreadsheet application, cached values may be absent (a known
    openpyxl/Excel-format limitation, not something this reader works
    around).

    open() infers the schema from up to _SCHEMA_SAMPLE_ROWS rows, not the
    whole file. read_batches() then streams the rest of the sheet,
    casting every batch to that same schema so the stream has one
    consistent dtype per column throughout, like every other reader.
    """

    def __init__(self, path: str | Path) -> None:
        self._path = Path(path)
        self._workbook: openpyxl.Workbook | None = None
        self._header: list[str] | None = None
        self._schema: Schema | None = None
        self._sample_rows: list[tuple[object, ...]] = []
        self._row_iter: Iterator[tuple[object, ...]] | None = None

    def open(self) -> None:
        """Validate the file exists and establish the schema from a
        bounded sample of rows, not the whole sheet."""
        if not self._path.is_file():
            raise FileNotFoundError(f"Excel file not found: {self._path}")

        self._workbook = openpyxl.load_workbook(
            self._path, read_only=True, data_only=True
        )
        sheet = self._workbook.active
        row_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration:
            raise ValueError(f"Excel file has no header row: {self._path}") from None
        self._header = [str(cell) if cell is not None else "" for cell in header_row]

        sample: list[tuple[object, ...]] = []
        for row in row_iter:
            sample.append(row)
            if len(sample) >= _SCHEMA_SAMPLE_ROWS:
                break

        if sample:
            inferred = pl.DataFrame(sample, schema=self._header, orient="row")
            self._schema = Schema(columns=dict(inferred.schema))
        else:
            # No data rows to infer from (header-only sheet): fall back to
            # treating every column as String rather than guessing.
            self._schema = Schema(columns=dict.fromkeys(self._header, pl.String()))
        self._sample_rows = sample
        self._row_iter = row_iter

    def schema(self) -> Schema:
        """Return the dataset schema. Valid only after open()."""
        if self._schema is None:
            raise RuntimeError("ExcelReader.schema() called before open()")
        return self._schema

    def read_batches(self, rows_per_batch: int) -> Iterator[pl.DataFrame]:
        """Yield row batches of at most rows_per_batch rows, streaming
        from openpyxl's read-only row iterator."""
        if self._row_iter is None:
            raise RuntimeError("ExcelReader.read_batches() called before open()")

        # Chain the already-consumed sample rows back onto the rest of the
        # iterator so every batch is sliced at exact rows_per_batch
        # boundaries, regardless of how the sample size relates to it -
        # the sample and the rest of the sheet are just one row stream.
        all_rows = itertools.chain(self._sample_rows, self._row_iter)
        self._sample_rows = []
        while True:
            batch_rows = list(itertools.islice(all_rows, rows_per_batch))
            if not batch_rows:
                return
            yield self._rows_to_frame(batch_rows)

    def _rows_to_frame(self, rows: list[tuple[object, ...]]) -> pl.DataFrame:
        assert self._header is not None and self._schema is not None
        return pl.DataFrame(rows, schema=self._schema.columns, orient="row")

    def close(self) -> None:
        """Release the file handle. Idempotent."""
        if self._workbook is not None:
            self._workbook.close()
        self._workbook = None
        self._header = None
        self._schema = None
        self._sample_rows = []
        self._row_iter = None
